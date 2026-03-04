"""
results_to_excel.py — Convert scraper results JSON to a formatted Excel spreadsheet

Usage:
  python results_to_excel.py dearborn_mi_results.json
  python results_to_excel.py dearborn_mi_results.json --output my_leads.xlsx

The Excel file contains two sheets:
  1. "Leads"   — one row per restaurant, all fields, colour-coded by ICP score
  2. "Summary" — stats: total leads, avg score, score distribution, top menu signals
"""

import argparse
import json
import os
import re
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1A1A2E"   # dark navy
C_HEADER_FG   = "FFFFFF"   # white
C_ACCENT      = "E94560"   # red-pink accent
C_SCORE_HIGH  = "C6EFCE"   # green fill  (score 8-10)
C_SCORE_MED   = "FFEB9C"   # yellow fill (score 6-7)
C_SCORE_LOW   = "FFC7CE"   # red fill    (score 1-5)
C_SCORE_HIGH_F = "276221"
C_SCORE_MED_F  = "9C6500"
C_SCORE_LOW_F  = "9C0006"
C_ROW_ALT     = "F5F5F5"   # light grey alternating row
C_ROW_WHITE   = "FFFFFF"
C_BORDER      = "D0D0D0"


def hex_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def thin_border() -> Border:
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def score_fill(score: int):
    if score >= 8:
        return hex_fill(C_SCORE_HIGH), C_SCORE_HIGH_F
    elif score >= 6:
        return hex_fill(C_SCORE_MED), C_SCORE_MED_F
    else:
        return hex_fill(C_SCORE_LOW), C_SCORE_LOW_F


def score_bar(score: int, max_score: int = 10) -> str:
    """Unicode bar representation of score."""
    filled = round(score / max_score * 8)
    return "█" * filled + "░" * (8 - filled)


# ── Sheet 1: Leads ────────────────────────────────────────────────────────────

COLUMNS = [
    ("Name",           28),
    ("Address",        36),
    ("Phone",          16),
    ("Website",        32),
    ("Email",          28),
    ("Rating ⭐",       10),
    ("Reviews 💬",      12),
    ("POS Platform",   14),
    ("ICP Score",      11),
    ("Busy Score",     11),
    ("Youth Score",    12),
    ("Menu Signals",   38),
    ("Reason",         48),
    ("Maps URL",       14),
]


def build_leads_sheet(ws, restaurants: list):
    ws.title = "Leads"
    ws.sheet_view.showGridLines = False

    # ── Title row ──
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"🍔  Restaurant Leads — ICP Matched  ({len(restaurants)} total)"
    title_cell.font = Font(name="Arial", bold=True, size=14, color=C_HEADER_FG)
    title_cell.fill = hex_fill(C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Generated date ──
    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:N2")
    date_cell = ws["A2"]
    date_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}"
    date_cell.font = Font(name="Arial", size=9, italic=True, color="888888")
    date_cell.fill = hex_fill("F0F0F0")
    date_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Column headers (row 3) ──
    ws.row_dimensions[3].height = 28
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, size=10, color=C_HEADER_FG)
        cell.fill = hex_fill(C_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = thin_border()

    # ── Freeze panes below header ──
    ws.freeze_panes = "A4"

    # ── Data rows ──
    for row_idx, r in enumerate(restaurants, 4):
        ws.row_dimensions[row_idx].height = 52

        pos         = r.get("pos", {})
        pos_name    = pos.get("platform", "Unknown")
        icp     = r.get("icp", {})
        score   = icp.get("score", 0)
        busy    = icp.get("busy_score", 0)
        youth   = icp.get("youth_score", 0)
        signals = ", ".join(icp.get("menu_signals", []))
        reason  = icp.get("reason", "")

        row_bg = C_ROW_ALT if row_idx % 2 == 0 else C_ROW_WHITE
        score_bg, score_fg = score_fill(score)

        values = [
            r.get("name", ""),
            r.get("address", ""),
            r.get("phone", ""),
            r.get("website", ""),
            r.get("email", ""),
            r.get("rating", ""),
            r.get("review_count", ""),
            pos_name,
            score,
            busy,
            youth,
            signals,
            reason,
            r.get("maps_url", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()

            # Score columns get colour coding
            if col_idx == 8:  # POS Platform
                # colour by known vs unknown
                if val and val != "Unknown":
                    cell.fill = hex_fill("D9EAD3")  # soft green
                    cell.font = Font(name="Arial", size=9, bold=True, color="274E13")
                else:
                    cell.fill = hex_fill(row_bg)
                    cell.font = Font(name="Arial", size=9, color="999999", italic=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 9:  # ICP Score
                cell.fill = score_bg
                cell.font = Font(name="Arial", size=10, bold=True, color=score_fg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (10, 11):  # Busy / Youth score
                cell.fill = score_bg
                cell.font = Font(name="Arial", size=9, color=score_fg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 6:  # Rating
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.fill = hex_fill(row_bg)
            elif col_idx == 7:  # Review count
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.number_format = "#,##0"
                cell.fill = hex_fill(row_bg)
            elif col_idx == 14:  # Maps URL
                if val:
                    cell.value = "Open ↗"
                    cell.hyperlink = val
                    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
                    cell.fill = hex_fill(row_bg)
            else:
                cell.fill = hex_fill(row_bg)

    # ── Auto-filter on header row ──
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNS))}{3 + len(restaurants)}"


# ── Sheet 2: Summary ──────────────────────────────────────────────────────────

def build_summary_sheet(ws, restaurants: list):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    def hdr(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Arial", bold=True, size=11, color=C_HEADER_FG)
        c.fill = hex_fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = thin_border()
        return c

    def lbl(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Arial", bold=True, size=10)
        c.alignment = Alignment(vertical="center", indent=1)
        c.border = thin_border()
        return c

    def val(row, col, value, fmt=None, color=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", size=10, color=color or "000000")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        if fmt:
            c.number_format = fmt
        return c

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    scores  = [r.get("icp", {}).get("score", 0) for r in restaurants]
    busys   = [r.get("icp", {}).get("busy_score", 0) for r in restaurants]
    youths  = [r.get("icp", {}).get("youth_score", 0) for r in restaurants]
    reviews = [r.get("review_count") or 0 for r in restaurants]
    ratings = [float(r.get("rating") or 0) for r in restaurants]

    avg = lambda lst: round(sum(lst) / len(lst), 1) if lst else 0

    # ── Title ──
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 36
    t = ws["A1"]
    t.value = "📊  Lead Generation Summary"
    t.font = Font(name="Arial", bold=True, size=14, color=C_HEADER_FG)
    t.fill = hex_fill(C_HEADER_BG)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Key stats ──
    ws.row_dimensions[2].height = 6
    ws.row_dimensions[3].height = 26
    ws.merge_cells("A3:D3")
    hdr(3, 1, "  KEY METRICS")

    stat_rows = [
        ("Total ICP-Matched Leads",   len(restaurants), None,    "276221"),
        ("Average ICP Score",          avg(scores),      "0.0",   None),
        ("Average Busy Score",         avg(busys),       "0.0",   None),
        ("Average Youth Score",        avg(youths),      "0.0",   None),
        ("Average Google Rating",      avg(ratings),     "0.00",  None),
        ("Average Review Count",       avg(reviews),     "#,##0", None),
        ("Total Reviews (all leads)",  sum(reviews),     "#,##0", None),
    ]

    for i, (label, value, fmt, color) in enumerate(stat_rows, 4):
        ws.row_dimensions[i].height = 22
        lbl(i, 1, label)
        val(i, 2, value, fmt, color)
        # alternating row
        fill = hex_fill(C_ROW_ALT if i % 2 == 0 else C_ROW_WHITE)
        ws.cell(i, 1).fill = fill
        ws.cell(i, 2).fill = fill

    # ── Score distribution ──
    dist_start = 4 + len(stat_rows) + 2
    ws.row_dimensions[dist_start - 1].height = 8
    ws.row_dimensions[dist_start].height = 26
    ws.merge_cells(f"A{dist_start}:D{dist_start}")
    hdr(dist_start, 1, "  ICP SCORE DISTRIBUTION")

    bands = [("🟢 High (8–10)", 8, 10), ("🟡 Medium (6–7)", 6, 7), ("🔴 Low (1–5)", 1, 5)]
    for j, (label, lo, hi) in enumerate(bands, dist_start + 1):
        ws.row_dimensions[j].height = 22
        count = sum(1 for s in scores if lo <= s <= hi)
        pct   = count / len(scores) * 100 if scores else 0
        bar   = score_bar(count, len(scores)) if scores else ""
        lbl(j, 1, label)
        val(j, 2, count, "#,##0")
        val(j, 3, pct / 100, "0%")
        val(j, 4, bar)
        row_fill = hex_fill(C_SCORE_HIGH if lo >= 8 else C_SCORE_MED if lo >= 6 else C_SCORE_LOW)
        for c in range(1, 5):
            ws.cell(j, c).fill = row_fill

    # ── Top menu signals ──
    sig_start = dist_start + len(bands) + 3
    ws.row_dimensions[sig_start - 1].height = 8
    ws.row_dimensions[sig_start].height = 26
    ws.merge_cells(f"A{sig_start}:D{sig_start}")
    hdr(sig_start, 1, "  TOP MENU SIGNALS (Youth Indicators)")

    all_signals = []
    for r in restaurants:
        all_signals.extend(r.get("icp", {}).get("menu_signals", []))
    signal_counts = Counter(s.lower() for s in all_signals).most_common(12)

    lbl(sig_start + 1, 1, "Menu Item")
    lbl(sig_start + 1, 2, "Count")
    lbl(sig_start + 1, 3, "Frequency")
    ws.row_dimensions[sig_start + 1].height = 22
    for col in range(1, 4):
        ws.cell(sig_start + 1, col).fill = hex_fill(C_ACCENT)
        ws.cell(sig_start + 1, col).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")

    for k, (sig, count) in enumerate(signal_counts, sig_start + 2):
        ws.row_dimensions[k].height = 20
        pct = count / len(restaurants) * 100 if restaurants else 0
        fill = hex_fill(C_ROW_ALT if k % 2 == 0 else C_ROW_WHITE)
        lbl(k, 1, sig.title())
        val(k, 2, count, "#,##0")
        val(k, 3, pct / 100, "0%")
        for c in range(1, 4):
            ws.cell(k, c).fill = fill

    # ── Contact completeness ──
    has_web   = sum(1 for r in restaurants if r.get("website"))
    has_phone = sum(1 for r in restaurants if r.get("phone"))
    has_email = sum(1 for r in restaurants if r.get("email"))

    contact_start = sig_start + len(signal_counts) + 3
    ws.row_dimensions[contact_start].height = 26
    ws.merge_cells(f"A{contact_start}:D{contact_start}")
    hdr(contact_start, 1, "  CONTACT COMPLETENESS")

    contact_rows = [
        ("Have Website",  has_web),
        ("Have Phone",    has_phone),
        ("Have Email",    has_email),
    ]
    for m, (label, count) in enumerate(contact_rows, contact_start + 1):
        ws.row_dimensions[m].height = 22
        pct = count / len(restaurants) if restaurants else 0
        fill = hex_fill(C_ROW_ALT if m % 2 == 0 else C_ROW_WHITE)
        lbl(m, 1, label)
        val(m, 2, count, "#,##0")
        val(m, 3, pct, "0%")
        for c in range(1, 4):
            ws.cell(m, c).fill = fill

    # ── POS Platform breakdown ──
    pos_counts = Counter(
        r.get("pos", {}).get("platform", "Unknown") for r in restaurants
    )
    pos_start = contact_start + len(contact_rows) + 3
    ws.row_dimensions[pos_start].height = 26
    ws.merge_cells(f"A{pos_start}:D{pos_start}")
    hdr(pos_start, 1, "  POS / ORDERING PLATFORM DETECTED")

    lbl(pos_start + 1, 1, "Platform")
    lbl(pos_start + 1, 2, "Count")
    lbl(pos_start + 1, 3, "% of Leads")
    ws.row_dimensions[pos_start + 1].height = 22
    for col in range(1, 4):
        ws.cell(pos_start + 1, col).fill = hex_fill(C_ACCENT)
        ws.cell(pos_start + 1, col).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")

    for n, (platform, count) in enumerate(pos_counts.most_common(), pos_start + 2):
        ws.row_dimensions[n].height = 20
        pct = count / len(restaurants) if restaurants else 0
        fill = hex_fill("D9EAD3" if platform != "Unknown" else (C_ROW_ALT if n % 2 == 0 else C_ROW_WHITE))
        lbl(n, 1, platform)
        val(n, 2, count, "#,##0")
        val(n, 3, pct, "0%")
        for c in range(1, 4):
            ws.cell(n, c).fill = fill


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Convert scraper results JSON to a formatted Excel spreadsheet",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python results_to_excel.py dearborn_mi_results.json
  python results_to_excel.py dearborn_mi_results.json --output my_leads.xlsx
        """
    )
    parser.add_argument("input", help="Path to the _results.json file from the scraper")
    parser.add_argument("--output", default=None,
                        help="Output .xlsx filename (default: same name as input with .xlsx)")
    args = parser.parse_args()

    # Load data
    if not os.path.exists(args.input):
        print(f"❌ File not found: {args.input}")
        return

    with open(args.input, "r", encoding="utf-8") as f:
        restaurants = json.load(f)

    if not restaurants:
        print("⚠  No restaurants found in results file.")
        return

    # Determine output path
    output_path = args.output
    if not output_path:
        base = os.path.splitext(args.input)[0]
        output_path = base + ".xlsx"

    print(f"\n📊 Building Excel workbook...")
    print(f"   Input  : {args.input} ({len(restaurants)} leads)")
    print(f"   Output : {output_path}")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Leads
    leads_ws = wb.create_sheet("Leads")
    build_leads_sheet(leads_ws, restaurants)
    print(f"   ✅ 'Leads' sheet — {len(restaurants)} rows")

    # Sheet 2: Summary
    summary_ws = wb.create_sheet("Summary")
    build_summary_sheet(summary_ws, restaurants)
    print(f"   ✅ 'Summary' sheet")

    # Set Leads as the active (first-seen) sheet
    wb.active = leads_ws

    wb.save(output_path)
    print(f"\n✅ Done! Saved to: {output_path}\n")


if __name__ == "__main__":
    main()