"""
filter_nearby_pos.py
────────────────────
Reads a results.json file from the scraper, then filters to restaurants that:
  1. Are within a max driving time from a given origin address
  (All restaurants are included regardless of detected POS platform)

Outputs a formatted Excel file with one sheet per filter + a summary.

Usage:
  python filter_nearby_pos.py results.json --origin "7625 E Morrow Circle, Dearborn MI"
  python filter_nearby_pos.py results.json --origin "7625 E Morrow Circle" --max-minutes 30
  python filter_nearby_pos.py results.json --origin "7625 E Morrow Circle" --max-minutes 20 --output nearby_leads.xlsx

Requirements:
  pip install requests openpyxl
  A Google Maps API key with Distance Matrix API enabled.
  Set env var: GOOGLE_MAPS_KEY=your_key
  OR pass via: --google-key YOUR_KEY
"""

import argparse
import json
import os
import re
import time
from collections import Counter
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Colours ───────────────────────────────────────────────────────────────────
C_NAVY     = "1A1A2E"
C_ACCENT   = "E94560"
C_WHITE    = "FFFFFF"
C_ALT      = "F5F5F5"
C_GREEN_BG = "C6EFCE"
C_GREEN_FG = "276221"
C_YEL_BG   = "FFEB9C"
C_YEL_FG   = "9C6500"
C_RED_BG   = "FFC7CE"
C_RED_FG   = "9C0006"
C_POS_BG   = "D9EAD3"
C_POS_FG   = "274E13"
C_BORDER   = "D0D0D0"


def hfill(color):
    return PatternFill("solid", start_color=color, fgColor=color)


def tborder():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def score_colors(score):
    if score >= 8:   return hfill(C_GREEN_BG), C_GREEN_FG
    if score >= 6:   return hfill(C_YEL_BG),   C_YEL_FG
    return               hfill(C_RED_BG),   C_RED_FG


def drive_color(minutes):
    """Green ≤15 min, yellow 15-25, red >25."""
    if minutes <= 15: return hfill(C_GREEN_BG), C_GREEN_FG
    if minutes <= 25: return hfill(C_YEL_BG),   C_YEL_FG
    return                   hfill(C_RED_BG),   C_RED_FG


# ── Coordinate extraction from Maps URL ───────────────────────────────────────

def extract_coords_from_url(maps_url: str):
    """
    Pull lat/lng from a Google Maps URL.
    Handles patterns like:
      /maps/place/Name/@37.123,-122.456,15z/
      /maps/place/Name/data=...!3d37.123!4d-122.456
    Returns (lat, lng) as floats, or (None, None).
    """
    if not maps_url:
        return None, None

    # Pattern 1: /@lat,lng,zoom
    m = re.search(r'/@(-?\d+\.\d+),(-?\d+\.\d+)', maps_url)
    if m:
        return float(m.group(1)), float(m.group(2))

    # Pattern 2: !3d<lat>!4d<lng>
    lat_m = re.search(r'!3d(-?\d+\.\d+)', maps_url)
    lng_m = re.search(r'!4d(-?\d+\.\d+)', maps_url)
    if lat_m and lng_m:
        return float(lat_m.group(1)), float(lng_m.group(1))

    return None, None


# ── Google Distance Matrix ────────────────────────────────────────────────────

def get_driving_minutes(origin: str, dest_lat: float, dest_lng: float, api_key: str) -> int | None:
    """
    Call Google Distance Matrix API.
    Returns driving time in minutes, or None on failure.
    """
    url = "https://maps.googleapis.com/maps/api/distancematrix/json"
    params = {
        "origins":      origin,
        "destinations": f"{dest_lat},{dest_lng}",
        "mode":         "driving",
        "key":          api_key,
    }
    try:
        resp = requests.get(url, params=params, timeout=10)
        data = resp.json()
        element = data["rows"][0]["elements"][0]
        if element["status"] == "OK":
            return round(element["duration"]["value"] / 60)   # seconds → minutes
        return None
    except Exception as e:
        print(f"      ⚠  Distance API error: {e}")
        return None


def geocode_address(origin: str, api_key: str):
    """Geocode origin address to lat/lng for fallback coord display."""
    url = "https://maps.googleapis.com/maps/api/geocode/json"
    try:
        resp = requests.get(url, params={"address": origin, "key": api_key}, timeout=10)
        loc = resp.json()["results"][0]["geometry"]["location"]
        return loc["lat"], loc["lng"]
    except Exception:
        return None, None


# ── Excel builder ─────────────────────────────────────────────────────────────

COLUMNS = [
    ("Name",           26),
    ("Address",        34),
    ("Phone",          16),
    ("Website",        30),
    ("Email",          26),
    ("POS Platform",   14),
    ("Drive Time",     12),
    ("Rating ⭐",       10),
    ("Reviews 💬",      12),
    ("ICP Score",      11),
    ("Busy Score",     11),
    ("Youth Score",    12),
    ("Menu Signals",   36),
    ("Maps URL",       12),
]


def write_header_block(ws, title: str, subtitle: str, col_count: int):
    end_col = get_column_letter(col_count)

    ws.row_dimensions[1].height = 38
    ws.merge_cells(f"A1:{end_col}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(name="Arial", bold=True, size=14, color=C_WHITE)
    c.fill = hfill(C_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[2].height = 16
    ws.merge_cells(f"A2:{end_col}2")
    c = ws["A2"]
    c.value = subtitle
    c.font = Font(name="Arial", size=9, italic=True, color="888888")
    c.fill = hfill("F0F0F0")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[3].height = 28
    for ci, (header, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        cell = ws.cell(row=3, column=ci, value=header)
        cell.font = Font(name="Arial", bold=True, size=10, color=C_WHITE)
        cell.fill = hfill(C_ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = tborder()

    ws.freeze_panes = "A4"


def write_data_rows(ws, restaurants: list, start_row: int = 4):
    for ri, r in enumerate(restaurants, start_row):
        ws.row_dimensions[ri].height = 50

        icp        = r.get("icp", {})
        pos        = r.get("pos", {})
        score      = icp.get("score", 0)
        busy       = icp.get("busy_score", 0)
        youth      = icp.get("youth_score", 0)
        signals    = ", ".join(icp.get("menu_signals", []))
        pos_name   = pos.get("platform", "Unknown")
        drive_mins = r.get("drive_minutes")

        row_bg = C_ALT if ri % 2 == 0 else C_WHITE
        s_fill, s_fg = score_colors(score)

        values = [
            r.get("name", ""),
            r.get("address", ""),
            r.get("phone", ""),
            r.get("website", ""),
            r.get("email", ""),
            pos_name,
            f"{drive_mins} min" if drive_mins is not None else "—",
            r.get("rating", ""),
            r.get("review_count", ""),
            score,
            busy,
            youth,
            signals,
            r.get("maps_url", ""),
        ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = tborder()

            if ci == 6:  # POS Platform
                cell.fill = hfill(C_POS_BG) if pos_name != "Unknown" else hfill(row_bg)
                cell.font = Font(name="Arial", size=9,
                                 bold=(pos_name != "Unknown"),
                                 color=C_POS_FG if pos_name != "Unknown" else "999999")
                cell.alignment = Alignment(horizontal="center", vertical="top")

            elif ci == 7:  # Drive Time
                if drive_mins is not None:
                    d_fill, d_fg = drive_color(drive_mins)
                    cell.fill = d_fill
                    cell.font = Font(name="Arial", size=10, bold=True, color=d_fg)
                else:
                    cell.fill = hfill(row_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif ci == 10:  # ICP Score
                cell.fill = s_fill
                cell.font = Font(name="Arial", size=10, bold=True, color=s_fg)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif ci in (11, 12):  # Busy / Youth
                cell.fill = s_fill
                cell.font = Font(name="Arial", size=9, color=s_fg)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif ci in (8, 9):  # Rating, Reviews
                cell.fill = hfill(row_bg)
                cell.alignment = Alignment(horizontal="center", vertical="top")
                if ci == 9:
                    cell.number_format = "#,##0"

            elif ci == 14:  # Maps URL → hyperlink
                if val:
                    cell.value = "Open ↗"
                    cell.hyperlink = val
                    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
                cell.fill = hfill(row_bg)

            else:
                cell.fill = hfill(row_bg)

    # Auto-filter
    last_row = start_row + len(restaurants) - 1
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNS))}{last_row}"


def build_leads_sheet(ws, restaurants: list, origin: str, max_minutes: int):
    ws.title = "Nearby Leads"
    ws.sheet_view.showGridLines = False

    subtitle = (
        f"Origin: {origin}  |  Max drive time: {max_minutes} min  |  "
        f"{len(restaurants)} leads  |  Generated: {datetime.now().strftime('%b %d, %Y %H:%M')}"
    )
    write_header_block(ws, "🍔  Nearby Restaurants — Qualified Leads", subtitle, len(COLUMNS))
    write_data_rows(ws, restaurants)


def build_summary_sheet(ws, all_restaurants: list, nearby: list, origin: str, max_minutes: int):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    def hdr(row, text, span="A:C"):
        cols = span.split(":")
        ws.merge_cells(f"{cols[0]}{row}:{cols[1]}{row}")
        c = ws.cell(row=row, column=1, value=f"  {text}")
        c.font = Font(name="Arial", bold=True, size=11, color=C_WHITE)
        c.fill = hfill(C_NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = tborder()
        ws.row_dimensions[row].height = 26

    def lbl(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Arial", bold=True, size=10)
        c.alignment = Alignment(vertical="center", indent=1)
        c.border = tborder()
        ws.row_dimensions[row].height = 22

    def val(row, col, value, fmt=None, color=None, bold=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", size=10, color=color or "000000", bold=bold)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = tborder()
        if fmt:
            c.number_format = fmt

    # Title
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 36
    t = ws["A1"]
    t.value = "📊  Filter Summary Report"
    t.font = Font(name="Arial", bold=True, size=14, color=C_WHITE)
    t.fill = hfill(C_NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.row_dimensions[2].height = 8

    # ── Filter parameters ──
    hdr(3, "FILTER PARAMETERS")
    params = [
        ("Origin Address",    origin),
        ("Max Drive Time",    f"{max_minutes} minutes"),
        ("Generated",         datetime.now().strftime("%B %d, %Y  %H:%M")),
    ]
    for i, (label, value) in enumerate(params, 4):
        ws.row_dimensions[i].height = 22
        fill = hfill(C_ALT if i % 2 == 0 else C_WHITE)
        lbl(i, 1, label)
        c = ws.cell(row=i, column=2, value=value)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = tborder()
        ws.merge_cells(f"B{i}:C{i}")
        ws.cell(i, 1).fill = fill
        ws.cell(i, 2).fill = fill

    ws.row_dimensions[7].height = 8

    # ── Pipeline numbers ──
    hdr(8, "PIPELINE NUMBERS")
    drive_times = [r["drive_minutes"] for r in nearby if r.get("drive_minutes") is not None]
    avg_drive = round(sum(drive_times) / len(drive_times), 1) if drive_times else 0
    scores = [r.get("icp", {}).get("score", 0) for r in nearby]
    avg_score = round(sum(scores) / len(scores), 1) if scores else 0

    # Count how many nearby have a detected POS
    nearby_with_pos = sum(
        1 for r in nearby
        if r.get("pos", {}).get("platform", "Unknown") != "Unknown"
    )

    stats = [
        ("Total restaurants loaded",               len(all_restaurants), None),
        ("Within driving distance",                len(nearby),          C_GREEN_FG),
        ("Nearby with POS detected",               nearby_with_pos,      C_POS_FG),
        ("Avg drive time (qualified leads)",       f"{avg_drive} min",   None),
        ("Avg ICP score (qualified leads)",        f"{avg_score}/10",    None),
    ]
    for i, (label, value, color) in enumerate(stats, 9):
        ws.row_dimensions[i].height = 22
        fill = hfill(C_ALT if i % 2 == 0 else C_WHITE)
        lbl(i, 1, label)
        val(i, 2, value, color=color, bold=bool(color))
        ws.cell(i, 1).fill = fill
        ws.cell(i, 2).fill = fill
        ws.merge_cells(f"B{i}:C{i}")

    ws.row_dimensions[14].height = 8

    # ── Drive time breakdown ──
    hdr(15, "DRIVE TIME BREAKDOWN")
    lbl(16, 1, "Band")
    lbl(16, 2, "Count")
    lbl(16, 3, "% of Leads")
    ws.row_dimensions[16].height = 22
    for ci in range(1, 4):
        ws.cell(16, ci).fill = hfill(C_ACCENT)
        ws.cell(16, ci).font = Font(name="Arial", bold=True, size=10, color=C_WHITE)

    bands = [("🟢 ≤15 min", 0, 15), ("🟡 16–25 min", 16, 25), ("🔴 26–30 min", 26, 30)]
    for bi, (label, lo, hi) in enumerate(bands, 17):
        count = sum(1 for t in drive_times if lo <= t <= hi)
        pct   = count / len(nearby) if nearby else 0
        f = hfill(C_GREEN_BG if lo == 0 else C_YEL_BG if lo == 16 else C_RED_BG)
        ws.row_dimensions[bi].height = 22
        lbl(bi, 1, label)
        val(bi, 2, count, "#,##0")
        val(bi, 3, pct, "0%")
        for ci in range(1, 4):
            ws.cell(bi, ci).fill = f

    ws.row_dimensions[20].height = 8

    # ── POS platform breakdown ──
    hdr(21, "POS PLATFORM BREAKDOWN")
    lbl(22, 1, "Platform")
    lbl(22, 2, "Count")
    lbl(22, 3, "% of Leads")
    ws.row_dimensions[22].height = 22
    for ci in range(1, 4):
        ws.cell(22, ci).fill = hfill(C_ACCENT)
        ws.cell(22, ci).font = Font(name="Arial", bold=True, size=10, color=C_WHITE)

    pos_counts = Counter(r.get("pos", {}).get("platform", "Unknown") for r in nearby)
    for pi, (platform, count) in enumerate(pos_counts.most_common(), 23):
        pct  = count / len(nearby) if nearby else 0
        fill = hfill(C_POS_BG) if platform != "Unknown" else hfill(C_ALT)
        ws.row_dimensions[pi].height = 22
        lbl(pi, 1, platform)
        val(pi, 2, count, "#,##0")
        val(pi, 3, pct, "0%")
        for ci in range(1, 4):
            ws.cell(pi, ci).fill = fill


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Filter results.json → nearby restaurants → Excel (all POS included)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python filter_nearby_pos.py dearborn_mi_results.json \\
      --origin "7625 E Morrow Circle, Dearborn MI"

  python filter_nearby_pos.py dearborn_mi_results.json \\
      --origin "7625 E Morrow Circle, Dearborn MI" --max-minutes 20

  python filter_nearby_pos.py dearborn_mi_results.json \\
      --origin "7625 E Morrow Circle" --max-minutes 30 --output nearby_leads.xlsx

  # Supply Google Maps API key inline:
  python filter_nearby_pos.py results.json \\
      --origin "7625 E Morrow Circle" --google-key AIza...
        """
    )
    parser.add_argument("input",          help="Path to _results.json from scraper")
    parser.add_argument("--origin",       required=True, help="Origin address for driving distance")
    parser.add_argument("--max-minutes",  type=int, default=30, help="Max driving minutes (default: 30)")
    parser.add_argument("--google-key",   default=None, help="Google Maps API key (or set GOOGLE_MAPS_KEY env var)")
    parser.add_argument("--output",       default=None, help="Output .xlsx filename")
    args = parser.parse_args()

    # ── API key ──
    api_key = args.google_key or os.environ.get("GOOGLE_MAPS_KEY", "")
    if not api_key:
        print("❌  No Google Maps API key found.")
        print("    Set env var GOOGLE_MAPS_KEY=your_key  OR  pass --google-key YOUR_KEY")
        print("    Get a free key at: https://console.cloud.google.com/  (enable Distance Matrix API)")
        return

    # ── Load data ──
    if not os.path.exists(args.input):
        print(f"❌  File not found: {args.input}")
        return

    with open(args.input, "r", encoding="utf-8") as f:
        all_restaurants = json.load(f)

    print(f"\n📂 Loaded {len(all_restaurants)} restaurants from {args.input}")

    # ── Driving distance filter (all restaurants, no POS pre-filter) ──
    print(f"\n🚗 Checking driving distance from: {args.origin}")
    print(f"   Max: {args.max_minutes} minutes\n")

    nearby = []
    skipped_no_coords = 0
    skipped_too_far   = 0

    for i, r in enumerate(all_restaurants, 1):
        name = r.get("name", "Unknown")
        maps_url = r.get("maps_url", "")

        lat, lng = extract_coords_from_url(maps_url)

        if lat is None:
            print(f"  [{i:02d}] ⚠  {name} — no coords in Maps URL, skipping")
            skipped_no_coords += 1
            continue

        minutes = get_driving_minutes(args.origin, lat, lng, api_key)

        if minutes is None:
            print(f"  [{i:02d}] ⚠  {name} — distance lookup failed, skipping")
            skipped_no_coords += 1
            continue

        pos_name = r.get("pos", {}).get("platform", "Unknown")
        if minutes <= args.max_minutes:
            r["drive_minutes"] = minutes
            nearby.append(r)
            pos_label = pos_name if pos_name != "Unknown" else "no POS"
            print(f"  [{i:02d}] ✅  {name} | {pos_label} | 🚗 {minutes} min")
        else:
            skipped_too_far += 1
            pos_label = pos_name if pos_name != "Unknown" else "no POS"
            print(f"  [{i:02d}] ❌  {name} | {pos_label} | 🚗 {minutes} min (too far)")

        time.sleep(0.1)   # gentle rate limiting

    print(f"\n{'─'*55}")
    print(f"  ✅ Qualified (within distance): {len(nearby)}")
    print(f"  ❌ Too far:                     {skipped_too_far}")
    print(f"  ⚠  No coordinates / API error:  {skipped_no_coords}")
    print(f"{'─'*55}\n")

    if not nearby:
        print("⚠  No restaurants passed the distance filter. Try increasing --max-minutes.")
        return

    # Sort by drive time ascending
    nearby.sort(key=lambda r: r.get("drive_minutes", 999))

    # ── Step 3: Build Excel ──
    output_path = args.output
    if not output_path:
        base = os.path.splitext(args.input)[0]
        output_path = f"{base}_nearby_leads.xlsx"

    print(f"📊 Building Excel → {output_path}")

    wb = Workbook()
    wb.remove(wb.active)

    leads_ws = wb.create_sheet("Nearby Leads")
    build_leads_sheet(leads_ws, nearby, args.origin, args.max_minutes)

    summary_ws = wb.create_sheet("Summary")
    build_summary_sheet(summary_ws, all_restaurants, nearby, args.origin, args.max_minutes)

    wb.active = leads_ws
    wb.save(output_path)

    print(f"✅ Saved {len(nearby)} leads → {output_path}\n")

    # Print table
    print(f"{'#':<4} {'Name':<28} {'POS':<12} {'Drive':<8} {'ICP':<6} {'Address'}")
    print("─" * 85)
    for i, r in enumerate(nearby, 1):
        pos_label = r.get("pos", {}).get("platform", "Unknown")
        print(
            f"{i:<4} {r.get('name','?')[:27]:<28} "
            f"{pos_label[:11]:<12} "
            f"{r.get('drive_minutes', '?')} min   "
            f"{r.get('icp',{}).get('score',0)}/10   "
            f"{r.get('address','')[:30]}"
        )


if __name__ == "__main__":
    main()